# app.py
import os, json
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, session, jsonify
from models import db, Student, Question, Attempt, Answer

import pandas as pd
from openpyxl import load_workbook

APP_ROOT = os.path.abspath(os.path.dirname(__file__))

STUDENTS_XLSX = os.path.join(APP_ROOT, "students.xlsx")  # place your excel here

app = Flask(__name__)
app.secret_key = "change_this_secret"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///exam.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)

with app.app_context():
    db.create_all()

# ----------------- helper: import students from excel -----------------
# Replace the import_students_from_excel function with this:

def import_students_from_excel():
    if not os.path.exists(STUDENTS_XLSX):
        print("students.xlsx not found; skipping import.")
        return
    
    print(f"Reading Excel file from: {STUDENTS_XLSX}")
    
    # Read Excel WITHOUT converting everything to string first
    df = pd.read_excel(STUDENTS_XLSX)
    
    # DEBUG: Print column names
    print(f"Excel columns found: {list(df.columns)}")
    print(f"Total rows in Excel: {len(df)}")
    
    # Clean column names (remove spaces)
    df.columns = df.columns.str.strip()
    
    imported_count = 0
    for idx, row in df.iterrows():
        # Convert to string and strip, handling NaN values
        roll = str(row.get("roll_no", "")).strip()
        name = str(row.get("name", "")).strip()
        email = str(row.get("email", "")).strip()
        password = str(row.get("password", "")).strip()
        
        # Remove .0 from end if password was read as float
        if password.endswith('.0'):
            password = password[:-2]
        
        print(f"Row {idx}: roll='{roll}', name='{name}', password='{password}'")
        
        if not roll or roll.lower() == "nan" or roll == "":
            print(f"  -> Skipping empty roll_no")
            continue
        
        # only add if not exists
        existing = Student.query.filter_by(roll_no=roll).first()
        if not existing:
            s = Student(
                roll_no=roll,
                name=name,
                email=email,
                password=password
            )
            db.session.add(s)
            imported_count += 1
            print(f"  -> Added new student: {roll} with password: {password}")
        else:
            # Update existing student's password in case it changed
            existing.password = password
            print(f"  -> Updated student {roll} password")
    
    db.session.commit()
    print(f"Students imported: {imported_count} new students added.")

# ----------------- DEBUG ROUTE (ADMIN ONLY) -----------------
@app.route("/debug/students")
def debug_students():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    
    students = Student.query.all()
    result = f"<h2>Students in Database (Total: {len(students)})</h2>"
    result += "<table border='1' cellpadding='5'>"
    result += "<tr><th>ID</th><th>Roll No</th><th>Name</th><th>Email</th><th>Password</th></tr>"
    for s in students:
        result += f"<tr><td>{s.id}</td><td>'{s.roll_no}'</td><td>{s.name}</td><td>{s.email}</td><td>'{s.password}'</td></tr>"
    result += "</table>"
    result += "<br><a href='/admin/dashboard'>Back to Admin Dashboard</a>"
    return result

@app.route("/debug/clear_students")
def debug_clear_students():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    
    Student.query.delete()
    db.session.commit()
    return "All students deleted. <a href='/admin/dashboard'>Back to Dashboard</a> | <a href='/'>Restart app to reimport</a>"

# ----------------- ADMIN (simple hardcoded admin) -----------------
@app.route("/")
def index():
    return """
    <h2>Online Exam System</h2>
    <ul>
        <li><a href="/admin">Admin Login</a></li>
        <li><a href="/student">Student Login</a></li>
    </ul>
    """

@app.route("/admin", methods=["GET","POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        if username == "admin" and password == "admin123":
            session["admin"] = True
            return redirect(url_for("admin_dashboard"))
        return "Invalid admin credentials"
    return render_template("admin_login.html")

@app.route("/admin/dashboard")
def admin_dashboard():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    return render_template("admin_dashboard.html")

@app.route("/admin/add_question", methods=["GET","POST"])
def admin_add_question():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    if request.method == "POST":
        q = Question(
            text = request.form.get("text"),
            opt_a = request.form.get("opt_a"),
            opt_b = request.form.get("opt_b"),
            opt_c = request.form.get("opt_c"),
            opt_d = request.form.get("opt_d"),
            correct = request.form.get("correct").upper(),
            per_question_time = int(request.form.get("per_question_time", 30)),
            order_index = int(request.form.get("order_index", 0))
        )
        db.session.add(q)
        db.session.commit()
        return redirect(url_for("admin_view_questions"))
    return render_template("add_question.html")

@app.route("/admin/view_questions")
def admin_view_questions():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    qs = Question.query.order_by(Question.order_index.asc()).all()
    return render_template("view_questions.html", questions=qs)

@app.route("/admin/delete_question/<int:q_id>")
def admin_delete_question(q_id):
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    q = Question.query.get(q_id)
    if q:
        db.session.delete(q)
        db.session.commit()
    return redirect(url_for("admin_view_questions"))

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("admin_login"))

# ----------------- STUDENT LOGIN / DASHBOARD -----------------
@app.route("/student", methods=["GET","POST"])
def student_login():
    if request.method == "POST":
        roll = str(request.form.get("roll_no")).strip()
        pwd = str(request.form.get("password")).strip()

        student = Student.query.filter_by(roll_no=roll, password=pwd).first()
        if not student:
            return "Invalid credentials"

        # 🚫 CHECK IF STUDENT ALREADY WROTE THE EXAM
        previous_attempt = Attempt.query.filter_by(student_id=student.id).first()
        if previous_attempt:
            return render_template("already_attempted.html", student=student)

        # If first time → allow login
        session["student_id"] = student.id
        session["student_roll"] = student.roll_no
        session["student_name"] = student.name

        return redirect(url_for("student_dashboard"))

    return render_template("student_login.html")


@app.route("/student/dashboard")
def student_dashboard():
    if not session.get("student_id"):
        return redirect(url_for("student_login"))
    return render_template("student_dashboard.html", name=session.get("student_name"))

# ----------------- START EXAM -----------------
@app.route("/student/start")
def student_start():
    if not session.get("student_id"):
        return redirect(url_for("student_login"))
    # create attempt
    attempt = Attempt(student_id=session["student_id"])
    db.session.add(attempt)
    db.session.commit()
    return render_template("exam.html", attempt_id=attempt.id)

@app.route("/api/questions/<int:attempt_id>")
def api_questions(attempt_id):
    # return questions ordered with per_question_time
    qs = Question.query.order_by(Question.order_index.asc()).all()
    payload = []
    total_time = 0
    for q in qs:
        payload.append({
            "id": q.id,
            "text": q.text,
            "options": {"A": q.opt_a, "B": q.opt_b, "C": q.opt_c, "D": q.opt_d},
            "per_question_time": q.per_question_time
        })
        total_time += q.per_question_time
    return jsonify({"questions": payload, "total_time": total_time})

# ----------------- SUBMIT ANSWER / FINISH -----------------
@app.route("/api/submit_answers", methods=["POST"])
def api_submit_answers():
    data = request.json
    attempt_id = data.get("attempt_id")
    answers = data.get("answers", {})  # {question_id: "A", ...}
    # find attempt
    att = Attempt.query.get(attempt_id)
    if not att:
        return jsonify({"status":"error", "msg":"Attempt not found"}), 404
    # save answers
    for qid_str, sel in answers.items():
        try:
            qid = int(qid_str)
        except:
            continue
        ans = Answer(attempt_id=attempt_id, question_id=qid, selected=sel)
        db.session.add(ans)
    att.finished_at = datetime.utcnow()
    # score
    db.session.flush()
    total_score = 0
    for a in Answer.query.filter_by(attempt_id=attempt_id).all():
        q = Question.query.get(a.question_id)
        if q and a.selected and a.selected.upper() == (q.correct or "").upper():
            total_score += 1
    att.score = total_score
    db.session.add(att)
    db.session.commit()
    # write results back to excel
    try:
        write_results_to_excel(attempt_id)
    except Exception as e:
        print("Excel write error:", e)
    return jsonify({"status": "ok"})

# ----------------- ADMIN RESULTS PAGE -----------------
@app.route("/admin/results")
def admin_results():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    
    attempts = Attempt.query.order_by(Attempt.id.desc()).all()
    results = []

    for att in attempts:
        student = Student.query.get(att.student_id)
        duration = None
        if att.started_at and att.finished_at:
            duration = att.finished_at - att.started_at

        results.append({
            "attempt_id": att.id,
            "roll_no": student.roll_no,
            "name": student.name,
            "score": att.score,
            "started_at": att.started_at,
            "finished_at": att.finished_at,
            "duration": duration
        })

    return render_template("results.html", results=results)


@app.route("/admin/results/<int:attempt_id>")
def admin_view_result(attempt_id):
    if not session.get("admin"):
        return redirect(url_for("admin_login"))

    att = Attempt.query.get(attempt_id)
    if not att:
        return "Attempt not found"

    student = Student.query.get(att.student_id)
    answers = Answer.query.filter_by(attempt_id=attempt_id).all()

    detailed = []

    for ans in answers:
        q = Question.query.get(ans.question_id)
        detailed.append({
            "qno": ans.question_id,
            "question": q.text,
            "correct": q.correct,
            "selected": ans.selected,
            "status": "Correct" if ans.selected == q.correct else ("Skipped" if ans.selected == "" else "Wrong")
        })

    return render_template("result_detail.html", student=student, detailed=detailed, attempt=att)


# ----------------- Write results to same Excel -----------------
def write_results_to_excel(attempt_id):
    """
    Writes attempt results into students.xlsx sheet 'answers'.
    Columns: roll_no, name, email, attempt_id, started_at, finished_at, Q_<id>..., score, timestamp
    """
    att = Attempt.query.get(attempt_id)
    if not att:
        return
    student = Student.query.get(att.student_id)
    answers = Answer.query.filter_by(attempt_id=attempt_id).all()
    # Build row dict
    row = {
        "roll_no": student.roll_no,
        "name": student.name,
        "email": student.email,
        "attempt_id": att.id,
        "started_at": att.started_at,
        "finished_at": att.finished_at,
        "score": att.score,
        "timestamp": datetime.utcnow()
    }
    # add each question column
    for a in answers:
        col = f"Q_{a.question_id}"
        row[col] = a.selected

    # Read existing excel
    if not os.path.exists(STUDENTS_XLSX):
        # create a new workbook if not exists
        df = pd.DataFrame([row])
        df.to_excel(STUDENTS_XLSX, sheet_name="answers", index=False)
        return

    # load workbook and existing answers sheet if present
    book = load_workbook(STUDENTS_XLSX)
    with pd.ExcelWriter(STUDENTS_XLSX, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        writer.book = book
        # read existing answers sheet if exists
        if "answers" in book.sheetnames:
            existing = pd.read_excel(STUDENTS_XLSX, sheet_name="answers")
            # append new row
            newdf = existing.append(row, ignore_index=True)
            # remove old answers sheet
            idx = book.sheetnames.index("answers")
            std = book["answers"]
            book.remove(std)
            writer.book = book
            newdf.to_excel(writer, sheet_name="answers", index=False)
        else:
            pd.DataFrame([row]).to_excel(writer, sheet_name="answers", index=False)

    print("Wrote results to Excel (answers sheet).")

# ----------------- Run app -----------------
if __name__ == "__main__":
    app.run(debug=True)